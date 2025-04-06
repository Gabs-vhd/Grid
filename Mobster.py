import tkinter as tk
from tkinter import filedialog
import openpyxl
import pygame
from pathlib import Path
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import random
import re
import pandas as pd
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import matplotlib.pyplot as plt
from collections import Counter

class Mobster(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Mobster")
        self.geometry("755x600")
        self.resizable(False,False)
        self.frames = {}
        self.current_page = 0
        
        # Caminho relativo para o ícone
        icon_path = os.path.join(os.path.dirname(__file__),"imagens", "skull_icone.ico")
        self.iconbitmap(icon_path)
        
        # configs grid principal
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        
        # Inicializar o pygame
        pygame.mixer.init()
        
        # Música de fundo
        self.bg_music_path = Path(__file__).parent / "imagens/bgmusic.mp3"
        self.page_sound_path = Path(__file__).parent / "imagens/pagetur.mp3"
        
        # Carregando os sons
        if self.bg_music_path.exists():
            pygame.mixer.music.load(str(self.bg_music_path))
            pygame.mixer.music.play(-1)
        
        self.page_sound = None
        if self.page_sound_path.exists():
            self.page_sound = pygame.mixer.Sound(str(self.page_sound_path))
        
        self.som_ativo = True
        
        # Páginas
        self.pages = [
            Page1(self),
            Page2(self),
            Page3(self),
            Page4(self),
            Page5(self),
            Page6(self, self.frames),
        ]
        
        for i, page in enumerate(self.pages):
            self.frames[i] = page
            page.grid(row=0, column=0, sticky="nsew")

        self.show_page(self.current_page)
        
        # Navegação
        navigation_frame = tk.Frame(self, background="#F5F5DC")
        navigation_frame.grid(row=1, column=0, sticky="ew")
        navigation_frame.grid_columnconfigure((0, 1), weight=1)

        btn_back = tk.Button(navigation_frame, text="Back", font=("Copperplate Gothic Bold", 20), bg="#FFA07A", border=3, command=self.previous_page)
        btn_back.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        btn_back.bind('<Button-1>', self.page_turning)

        btn_next = tk.Button(navigation_frame, text="Next", font=("Copperplate Gothic Bold", 20), bg="#4682B4", border=3, command=self.next_page)
        btn_next.grid(row=0, column=2, padx=10, pady=10, sticky="e")
        btn_next.bind('<Button-1>', self.page_turning)
        
        # Botão de som
        self.sound_button = tk.Button(navigation_frame,text="Sound On",font=("Copperplate Gothic Bold", 10),foreground="black",background="lightgrey", justify='left',border=2,width=10,command=self.toggle_sound)
        self.sound_button.grid(row=0, column=1, ipadx=8, ipady=8, padx=(20,120))
        
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        
    def toggle_sound(self):
        if self.som_ativo:
            pygame.mixer.music.pause()
            self.som_ativo = False
            self.sound_button.config(text="Sound Off", background="lightcoral")
        else:
            pygame.mixer.music.unpause()
            self.som_ativo = True
            self.sound_button.config(text="Sound On", background="lightgrey")
    
    def on_close(self):
        # Parar música e encerrar mixer
        pygame.mixer.music.stop()
        pygame.mixer.quit()
    
        # Fechar a janela
        self.destroy()
    
    def page_turning(self, event):
        if self.page_sound:
            self.page_sound.play()
    
    def show_page(self, page_number):
        for frame in self.frames.values():
            frame.grid_remove()
        self.frames[page_number].grid()

    def next_page(self):
        if self.current_page < len(self.frames) - 1:
            self.current_page += 1
            self.show_page(self.current_page)

    def previous_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.show_page(self.current_page)




class Page1(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.configure(background="#F8F8FF")
        
        self.autor = "Made By @gabs.vhd"

        # Variáveis
        self.battle_name_var = tk.StringVar()
        self.mob_name_var = tk.StringVar()
        self.quantity_var = tk.IntVar()

        # Carregar imagens
        skull_left = Path(__file__).parent / "imagens/skullL.png"
        skull_right = Path(__file__).parent / "imagens/skullR.png"

        if skull_left.exists():
            self.skull_image_left = tk.PhotoImage(file=str(skull_left))
        else:
            raise FileNotFoundError(f"Imagem '{skull_left}' não encontrada!")

        if skull_right.exists():
            self.skull_image_right = tk.PhotoImage(file=str(skull_right))
        else:
            raise FileNotFoundError(f"Imagem '{skull_right}' não encontrada!")

        # Canvas para fundo personalizado
        self.canvas = tk.Canvas(self, bg="#F5F5DC", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)

        # Adicionar fundo
        self.canvas.create_image(65, 60, anchor="nw", image=self.skull_image_left)
        self.canvas.create_image(725, 60, anchor="ne", image=self.skull_image_right)

        # Texto explicativo
        explanation_text = (
            "Choose a name like 'Troll' or 'Kobold'. Don't worry, "
            "they will be automatically numbered ('Kobold 1', 'Kobold 2', etc.)"
        )
        self.canvas.create_text(
            430, 50, text=explanation_text, font=("Copperplate Gothic Light", 10),
            fill="black", width=400, justify="center"
        )
        
        self.canvas.create_text(70,10, text=self.autor, font=("Copperplate Gothic Light", 10), fill="black", width=200, justify='left')

        # Inputs e labels
        self.create_label_and_entry(100, "Battle Name", self.battle_name_var, validate_fn=self.validate_filename)
        self.create_label_and_entry(200, "Mob Name", self.mob_name_var)
        self.create_label_and_entry(300, "Amount", self.quantity_var, is_numeric=True)

    def create_label_and_entry(self, y, label_text, variable, is_numeric=False, validate_fn=None):
        """
        Cria um rótulo e um campo de entrada alinhados no Canvas.
        - y: posição vertical
        - label_text: texto do rótulo
        - variable: variável associada ao campo de entrada
        - is_numeric: se True, o campo aceita apenas números
        - validate_fn: função de validação personalizada
        """
        # Adicionar texto no Canvas
        self.canvas.create_text(
            430, y, text=label_text, font=("Copperplate Gothic Bold", 20), fill="black"
        )

        # Criar entrada
        entry = tk.Entry(
            self, textvariable=variable, font=("Copperplate Gothic Light", 15),
            justify="left", width=20, border=3
        )
        self.canvas.create_window(430, y + 40, window=entry)

        # Configurar validação
        if is_numeric:
            entry.config(validate="key", validatecommand=(self.register(self.validate_numeric), "%P"))
        elif validate_fn:
            entry.config(validate="key", validatecommand=(self.register(validate_fn), "%P"))

    def validate_numeric(self, value):
        """Permite apenas valores numéricos nos campos."""
        return value.isdigit() or value == ""

    def validate_filename(self, value):
        """
        Valida se a entrada é um nome de arquivo válido.
        - value: valor digitado
        Retorna True se for válido, False caso contrário.
        """
        # Lista de caracteres proibidos para nomes de arquivo
        prohibited_characters = ['\\', '/', ':', '*', '?', '"', '<', '>', '|', '.']

        # Verifica se o valor contém algum caractere proibido
        if any(char in value for char in prohibited_characters):
            return False

        # Verifica se o valor é vazio ou contém apenas espaços
        if value.strip() == "":
            return True

        return True


class Page2(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.configure(background="#F5F5DC")

        # Variáveis
        self.hp_min_var = tk.IntVar(value=0)
        self.hp_max_var = tk.IntVar(value=0)
        self.defe_min_var = tk.IntVar(value=0)
        self.defe_max_var = tk.IntVar(value=0)

        # Carregar imagens
        shield_path = Path(__file__).parent / "imagens/shield.png"
        heart_path = Path(__file__).parent / "imagens/potion.png"

        if shield_path.exists():
            self.shield_image = tk.PhotoImage(file=str(shield_path))
        else:
            raise FileNotFoundError(f"Imagem '{shield_path}' não encontrada!")

        if heart_path.exists():
            self.heart_image = tk.PhotoImage(file=str(heart_path))
        else:
            raise FileNotFoundError(f"Imagem '{heart_path}' não encontrada!")

        # Canvas para adicionar imagens e texto
        self.canvas = tk.Canvas(self, bg="#F5F5DC", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)

        # Adicionar imagens ao Canvas
        self.canvas.create_image(150, 150, image=self.heart_image, anchor="center")
        self.canvas.create_image(600, 350, image=self.shield_image, anchor="center")

        # Texto explicativo
        explanation_text = (
            "Set the minimum and maximum values for hit points and defense points to randomize."
        )
        self.canvas.create_text(
            400, 50, text=explanation_text,
            font=("Copperplate Gothic Light", 12),
            fill="black", width=400, justify="center"
        )

        # Rótulo e inputs de HP
        self.create_label_and_input(150, "Hit Points", self.hp_min_var, self.hp_max_var)

        # Rótulo e inputs de Defesa
        self.create_label_and_input(300, "Defense (AC)", self.defe_min_var, self.defe_max_var)

        # Rótulo de erro
        self.error_label = tk.Label(self, text="", font=("Copperplate Gothic Light", 12), fg="red", bg="#F5F5DC")
        self.canvas.create_window(400, 450, window=self.error_label)

    def create_label_and_input(self, y, label_text, min_var, max_var):
        """
        Cria um rótulo com campos de entrada para min e max, usando o Canvas.
        - y: posição vertical no Canvas
        - label_text: texto do rótulo
        - min_var: variável associada ao campo "Min"
        - max_var: variável associada ao campo "Max"
        """
        self.canvas.create_text(
            400, y - 40, text=label_text,
            font=("Copperplate Gothic Bold", 25),
            fill="black"
        )

        # Texto "Min" e "Max"
        self.canvas.create_text(350, y, text="Min", font=("Copperplate Gothic Bold", 20), fill="black")
        self.canvas.create_text(450, y, text="Max", font=("Copperplate Gothic Bold", 20), fill="black")

        # Campos de entrada
        min_entry = tk.Entry(self, textvariable=min_var, font=("Copperplate Gothic Light", 15), width=5, border=5)
        max_entry = tk.Entry(self, textvariable=max_var, font=("Copperplate Gothic Light", 15), width=5, border=5)

        # Adicionar entradas ao Canvas
        self.canvas.create_window(350, y + 40, window=min_entry)
        self.canvas.create_window(450, y + 40, window=max_entry)

        # Validações
        min_entry.config(validate="key", validatecommand=(self.register(self.validate_quantity), "%P"))
        max_entry.config(validate="key", validatecommand=(self.register(self.validate_quantity), "%P"))

        # Adicionar validação de consistência
        max_entry.bind("<FocusOut>", lambda e: self.validate_min_max(label_text))

    def validate_quantity(self, value):
        """Permite apenas números inteiros no campo de entrada."""
        if value == "":
            return True
        pattern = r"^-?\d*$"
        return re.match(pattern, value) is not None

    def validate_min_max(self, field):
        """Valida se o valor máximo é maior ou igual ao mínimo."""
        if field == "Hit Points":
            min_value = self.hp_min_var.get()
            max_value = self.hp_max_var.get()
            if max_value < min_value:
                self.error_label.config(text="Hit Points: Maximum value must be bigger or equal than minimum!")
            else:
                self.error_label.config(text="")
        elif field == "Defense (CA)":
            min_value = self.defe_min_var.get()
            max_value = self.defe_max_var.get()
            if max_value < min_value:
                self.error_label.config(text="Defense: Maximum value must be bigger or equal than minimum!")
            else:
                self.error_label.config(text="")

  
        
    
class Page3(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.configure(background="#F5F5DC")
        
        self.texto = "Enter the die value and the modifier value.\nThe initiative value is calculated by adding these two."
        
        # Caminhos das imagens
        dice_path = Path(__file__).parent / "imagens/dice.png"
        
        # Carregar imagens
        if dice_path.exists():
            self.dice_image = tk.PhotoImage(file=str(dice_path))
        else:
            raise FileNotFoundError(f"Imagem '{dice_path}' não encontrada!")
        
        
        # Canvas para gerenciar imagens e texto
        self.canvas = tk.Canvas(self, bg="#F5F5DC", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)
        

        # Adicionar imagens ao lado do rótulo "rewards"
        self.canvas.create_image(670, 267, image=self.dice_image, anchor="center")

        # Variáveis
        self.for_min_var = tk.IntVar()
        self.for_max_var = tk.IntVar()
        
        self.des_min_var = tk.IntVar()
        self.des_max_var = tk.IntVar()
        
        self.con_min_var = tk.IntVar()
        self.con_max_var = tk.IntVar()
        
        self.int_min_var = tk.IntVar()
        self.int_max_var = tk.IntVar()
        
        self.sab_min_var = tk.IntVar()
        self.sab_max_var = tk.IntVar()
        
        self.car_min_var = tk.IntVar()
        self.car_max_var = tk.IntVar()
        
        self.dado_init_value = tk.IntVar()
        self.mod_init_value = tk.IntVar()
        
        # Rótulo atributos
        tk.Label(self,text="Attributes",font=("Copperplate Gothic Bold", 20),bg="#F5F5DC",wraplength=400,justify="center",).place(relx=0.1, rely=0.05, relwidth=0.8)
        
        # Min/Max left
        tk.Label(self,text="Min",font=("Copperplate Gothic Bold", 15),bg="#F5F5DC",wraplength=400,justify="left", anchor="w").place(relx=0.155, rely=0.15, relwidth=0.2)
        tk.Label(self,text="Max",font=("Copperplate Gothic Bold", 15),bg="#F5F5DC",wraplength=400,justify="left", anchor="w").place(relx=0.255, rely=0.15, relwidth=0.2)
        
        # Min/Max right
        tk.Label(self,text="Min",font=("Copperplate Gothic Bold", 15),bg="#F5F5DC",wraplength=400,justify="left", anchor="w").place(relx=0.585, rely=0.15, relwidth=0.2)
        tk.Label(self,text="Max",font=("Copperplate Gothic Bold", 15),bg="#F5F5DC",wraplength=400,justify="left", anchor="w").place(relx=0.685, rely=0.15, relwidth=0.2)
        
        """ FORÇA """
        tk.Label(self, text="Str:", font=("Copperplate Gothic Bold", 17), bg="#F5F5DC", justify='left', anchor='w').place(relx=0.07, rely=0.22, relwidth=0.4)
        self.for_min = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.for_min_var)
        self.for_min.place(relx=0.15, rely=0.22, relwidth=0.07)
        
        self.for_max = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.for_max_var)
        self.for_max.place(relx=0.25, rely=0.22, relwidth=0.07)
        
        self.for_min.bind("<FocusOut>", lambda e: self.validate_min_max("Str"))
        self.for_max.bind("<FocusOut>", lambda e: self.validate_min_max("Str"))
        
        self.for_min.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        self.for_max.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        """ DESTREZA """
        tk.Label(self, text="Dex:", font=("Copperplate Gothic Bold", 17), bg="#F5F5DC", justify='left', anchor='w').place(relx=0.07, rely=0.35, relwidth=0.4)
        self.des_min = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.des_min_var)
        self.des_min.place(relx=0.15, rely=0.35, relwidth=0.07)
        
        self.des_max = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.des_max_var)
        self.des_max.place(relx=0.25, rely=0.35, relwidth=0.07)
        
        self.des_min.bind("<FocusOut>", lambda e: self.validate_min_max("Dex"))
        self.des_max.bind("<FocusOut>", lambda e: self.validate_min_max("Dex"))
        
        self.des_min.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        self.des_max.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        """ CONSTITUIÇÃO """
        tk.Label(self, text="Con:", font=("Copperplate Gothic Bold", 17), bg="#F5F5DC", justify='left', anchor='w').place(relx=0.07, rely=0.48, relwidth=0.4)
        self.con_min = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.con_min_var)
        self.con_min.place(relx=0.16, rely=0.48, relwidth=0.07)
        
        self.con_max = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.con_max_var)
        self.con_max.place(relx=0.25, rely=0.48, relwidth=0.07)
        
        self.con_min.bind("<FocusOut>", lambda e: self.validate_min_max("Con"))
        self.con_max.bind("<FocusOut>", lambda e: self.validate_min_max("Con"))
        
        self.con_min.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        self.con_max.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        """ INTELIGÊNCIA """
        tk.Label(self, text="Int:", font=("Copperplate Gothic Bold", 17), bg="#F5F5DC", justify='left', anchor='w').place(relx=0.50, rely=0.22, relwidth=0.2)
        self.int_min = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.int_min_var)
        self.int_min.place(relx=0.58, rely=0.22, relwidth=0.07)
        
        self.int_max = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.int_max_var)
        self.int_max.place(relx=0.68, rely=0.22, relwidth=0.07)
        
        self.int_min.bind("<FocusOut>", lambda e: self.validate_min_max("Int"))
        self.int_max.bind("<FocusOut>", lambda e: self.validate_min_max("Int"))
        
        self.int_min.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        self.int_max.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        """ Sabedoria """
        tk.Label(self, text="Wis:", font=("Copperplate Gothic Bold", 17), bg="#F5F5DC", justify='left', anchor='w').place(relx=0.50, rely=0.35, relwidth=0.2)
        self.sab_min = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.sab_min_var)
        self.sab_min.place(relx=0.58, rely=0.35, relwidth=0.07)
        
        self.sab_max = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.sab_max_var)
        self.sab_max.place(relx=0.68, rely=0.35, relwidth=0.07)
        
        self.sab_min.bind("<FocusOut>", lambda e: self.validate_min_max("Wis"))
        self.sab_max.bind("<FocusOut>", lambda e: self.validate_min_max("Wis"))
        
        self.sab_min.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        self.sab_max.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        """ CARISMA """
        tk.Label(self, text="Cha:", font=("Copperplate Gothic Bold", 17), bg="#F5F5DC", justify='left', anchor='w').place(relx=0.50, rely=0.48, relwidth=0.2)
        self.car_min = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.car_min_var)
        self.car_min.place(relx=0.59, rely=0.48, relwidth=0.07)
        
        self.car_max = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.car_max_var)
        self.car_max.place(relx=0.68, rely=0.48, relwidth=0.07)
        
        self.car_min.bind("<FocusOut>", lambda e: self.validate_min_max("Cha"))
        self.car_max.bind("<FocusOut>", lambda e: self.validate_min_max("Cha"))
        
        self.car_min.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        self.car_max.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        
        """ DADOS E MODIFICADOR, INICIATIVA"""
        tk.Label(self, text="Initiative Die:", font=("Copperplate Gothic Bold", 17), bg="#F5F5DC", justify='left', anchor='w').place(relx=0.07, rely=0.65)
        self.dado_iniciativa = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.dado_init_value)
        self.dado_iniciativa.place(relx=0.32, rely=0.65, relwidth=0.07)
        self.dado_iniciativa.config(validate="key", validatecommand=(self.register(self.apenas_int_pos), "%P"))
        
        tk.Label(self, text="+", font=("Copperplate Gothic Bold", 20), bg="#F5F5DC", justify='left', anchor='w').place(relx=0.395, rely=0.645)
        
        tk.Label(self, text="Modifier:", font=("Copperplate Gothic Bold", 17), bg="#F5F5DC", justify='left', anchor='w').place(relx=0.43, rely=0.65)
        self.modificador_iniciativa = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.mod_init_value)
        self.modificador_iniciativa.place(relx=0.60, rely=0.65, relwidth=0.07)
        self.modificador_iniciativa.config(validate="key", validatecommand=(self.register(self.validate_quantity), "%P"))
        
        # Explicação da iniciativa
        tk.Label(self, text=self.texto, font=("Copperplate Gothic Light", 12), bg="#F5F5DC", justify='left', anchor='w').place(relx=0.07, rely=0.75)
        
        # Rótulo de erro
        self.error_label = tk.Label(self, text="", font=("Copperplate Gothic Light", 12), fg="red", bg="#F5F5DC")
        self.error_label.place(relx=0.1, rely=0.9, relwidth=0.8)
        
    
    def validate_min_max(self, field):
        if field == "Str":
            min_value = self.for_min_var.get()
            max_value = self.for_max_var.get()
            if max_value < min_value:
                self.error_label.config(text=f"STR: Maximum value must be bigger or equal than minimum!")
            else:
                self.error_label.config(text="")
        
        elif field == "Dex":
            min_value = self.des_min_var.get()
            max_value = self.des_max_var.get()
            if max_value < min_value:
                self.error_label.config(text=f"DEX: Maximum value must be bigger or equal than minimum!")
            else:
                self.error_label.config(text="")
        
        elif field == "Con":
            min_value = self.con_min_var.get()
            max_value = self.con_max_var.get()
            if max_value < min_value:
                self.error_label.config(text=f"CON: Maximum value must be bigger or equal than minimum!")
            else:
                self.error_label.config(text="")
                
        elif field == "Int":
            min_value = self.int_min_var.get()
            max_value = self.int_max_var.get()
            if max_value < min_value:
                self.error_label.config(text=f"INT: Maximum value must be bigger or equal than minimum!")
            else:
                self.error_label.config(text="")
                
        elif field == "Wis":
            min_value = self.sab_min_var.get()
            max_value = self.sab_max_var.get()
            if max_value < min_value:
                self.error_label.config(text=f"WIS: Maximum value must be bigger or equal than minimum!")
            else:
                self.error_label.config(text="")
                
        elif field == "Cha":
            min_value = self.car_min_var.get()
            max_value = self.car_max_var.get()
            if max_value < min_value:
                self.error_label.config(text=f"CHA: Maximum value must be bigger or equal than minimum!")
            else:
                self.error_label.config(text="")
            
                
    def validate_quantity(self, value):
        if value == "":
            return True
        pattern = r"^-?\d*$"
        return re.match(pattern, value) is not None
    
    def apenas_int_pos(self, value):
        return value.isdigit() or value == ""
    

class Page4(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.configure(background="#F5F5DC")

        # Caminhos das imagens
        axe1_path = Path(__file__).parent / "imagens/axe1.png"
        axe2_path = Path(__file__).parent / "imagens/axe2.png"

        # Carregar imagens
        if axe1_path.exists():
            self.axe1_image = tk.PhotoImage(file=str(axe1_path))
        else:
            raise FileNotFoundError(f"Imagem '{axe1_path}' não encontrada!")

        if axe2_path.exists():
            self.axe2_image = tk.PhotoImage(file=str(axe2_path))
        else:
            raise FileNotFoundError(f"Imagem '{axe2_path}' não encontrada!")

        # Canvas para gerenciar imagens e texto
        self.canvas = tk.Canvas(self, bg="#F5F5DC", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)

        # Adicionar imagens ao lado do rótulo "Weapons"
        self.canvas.create_image(100, 80, image=self.axe2_image, anchor="center")
        self.canvas.create_image(700, 80, image=self.axe1_image, anchor="center")

        # Rótulo Weapons
        self.canvas.create_text(
            400, 80, text="Weapons",
            font=("Copperplate Gothic Bold", 20),
            fill="black"
        )

        # Texto explicativo
        explanation_text = (
            "You can select a single weapon for everyone or enter a list of weapons, separated by commas, to be randomized. Example: Axe, Bow, Morning Star, Sword."
        )
        self.canvas.create_text(
            400, 150, text=explanation_text,
            font=("Copperplate Gothic Light", 12),
            fill="black", width=400, justify="center"
        )

        # Caixa de texto para entrada de armas
        self.weapons = tk.Text(self, height=5, font=("Copperplate Gothic Light", 15), border=3)
        self.canvas.create_window(400, 350, window=self.weapons, width=600, height=300)

        # Bind para validar quando a caixa de texto perder o foco
        self.weapons.bind("<FocusOut>", self.remove_consecutive_commas)

    def remove_consecutive_commas(self, event=None):
        """Remove vírgulas consecutivas do conteúdo do Text widget."""
        content = self.weapons.get("1.0", "end").strip()
        cleaned_content = ",".join(filter(None, content.split(",")))
        self.weapons.delete("1.0", "end")
        self.weapons.insert("1.0", cleaned_content)


class Page5(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.configure(background="#F5F5DC")
        
        # Variáveis
        self.gold_min_var = tk.IntVar()
        self.gold_max_var = tk.IntVar()
        
        self.xp_min_var = tk.IntVar()
        self.xp_max_var = tk.IntVar()
        
        # Caminhos das imagens
        coin_path = Path(__file__).parent / "imagens/coin.png"
        bread_path = Path(__file__).parent / "imagens/bread.png"
        
        # Carregar imagens
        if coin_path.exists():
            self.coin_image = tk.PhotoImage(file=str(coin_path))
        else:
            raise FileNotFoundError(f"Imagem '{coin_path}' não encontrada!")
        
        if bread_path.exists():
            self.bread_image = tk.PhotoImage(file=str(bread_path))
        else:
            raise FileNotFoundError(f"Imagem '{bread_path}' não encontrada!")
        
        # Canvas para gerenciar imagens e texto
        self.canvas = tk.Canvas(self, bg="#F5F5DC", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)
        

        # Adicionar imagens ao lado do rótulo "rewards"
        self.canvas.create_image(100, 200, image=self.coin_image, anchor="center")
        self.canvas.create_image(650, 220, image=self.bread_image, anchor="center")
        
        # Texto
        self.texto = "You can select a single item for everyone or enter a list of items, separated by commas, to be randomized. Example: Food, Tooth, Rope."
        
        # Rótulo Weapons
        tk.Label(self,text="Rewards",font=("Copperplate Gothic Bold", 20),bg="#F5F5DC",wraplength=400,justify="center").place(relx=0.1, rely=0.05, relwidth=0.8)
        
        # Rótulo Itens
        tk.Label(self,text="Itens",font=("Copperplate Gothic Bold", 20),bg="#F5F5DC",wraplength=400,justify="center").place(relx=0.44, rely=0.33, relwidth=0.1)
        
        tk.Label(self,text=self.texto,font=("Copperplate Gothic Light", 12),bg="#F5F5DC",wraplength=400,justify="center",).place(relx=0.22, rely=0.4, relwidth=0.55)
        
        # Min/Max left
        tk.Label(self,text="Min",font=("Copperplate Gothic Bold", 15),bg="#F5F5DC",wraplength=400,justify="left", anchor="w").place(relx=0.190, rely=0.15, relwidth=0.8)
        tk.Label(self,text="Max",font=("Copperplate Gothic Bold", 15),bg="#F5F5DC",wraplength=400,justify="left", anchor="w").place(relx=0.320, rely=0.15, relwidth=0.8)
        
        # Min/Max right
        tk.Label(self,text="Min",font=("Copperplate Gothic Bold", 15),bg="#F5F5DC",wraplength=400,justify="left", anchor="w").place(relx=0.61, rely=0.15, relwidth=0.8)
        tk.Label(self,text="Max",font=("Copperplate Gothic Bold", 15),bg="#F5F5DC",wraplength=400,justify="left", anchor="w").place(relx=0.741, rely=0.15, relwidth=0.8)
        
        """ GOLD"""
        tk.Label(self, text="Gold:", font=("Copperplate Gothic Bold", 17), bg="#F5F5DC", justify='left', anchor='w').place(relx=0.07, rely=0.22, relwidth=0.4)
        self.gold_min = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.gold_min_var)
        self.gold_min.place(relx=0.17, rely=0.22, relwidth=0.1)
        self.gold_min.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        self.gold_min.bind("<FocusOut>", lambda e: self.validate_min_max("Gold"))
        self.gold_min.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        self.gold_max = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.gold_max_var)
        self.gold_max.place(relx=0.30, rely=0.22, relwidth=0.1)
        self.gold_max.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        self.gold_max.bind("<FocusOut>", lambda e: self.validate_min_max("Gold"))
        self.gold_max.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        
        """ XP """
        tk.Label(self, text="XP:", font=("Copperplate Gothic Bold", 17), bg="#F5F5DC", justify='left', anchor='w').place(relx=0.50, rely=0.22, relwidth=0.4)
        self.xp_min = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.xp_min_var)
        self.xp_min.place(relx=0.58, rely=0.22, relwidth=0.110)
        self.xp_min.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        self.xp_min.bind("<FocusOut>", lambda e: self.validate_min_max("XP"))
        self.xp_min.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        self.xp_max = tk.Entry(self, font=("Copperplate Gothic Light", 12), border=3, textvariable=self.xp_max_var)
        self.xp_max.place(relx=0.72, rely=0.22, relwidth=0.110)
        self.xp_max.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        self.xp_max.bind("<FocusOut>", lambda e: self.validate_min_max("XP"))
        self.xp_max.config(validate="key",validatecommand=(self.register(self.validate_quantity), "%P"))
        
        """ TEXTO """
        self.itens = tk.Text(self, height=5, font=("Copperplate Gothic Light", 15), border=3)
        self.itens.place(relx=0.07, rely=0.52, relwidth=0.85, relheight=0.3)
        
        # Rótulo de erro
        self.error_label = tk.Label(self, text="", font=("Copperplate Gothic Light", 12), fg="red", bg="#F5F5DC")
        self.error_label.place(relx=0.1, rely=0.9, relwidth=0.8)
        
    def validate_quantity(self, value):
        if value == "":
            return True
        pattern = r"^-?\d*$"
        return re.match(pattern, value) is not None
    
    def validate_min_max(self, field):
        if field == "Gold":
            min_value = self.gold_min_var.get()
            max_value = self.gold_max_var.get()
            if max_value < min_value:
                self.error_label.config(text=f"Gold: Maximum value must be bigger or equal than minimum!")
            else:
                self.error_label.config(text="")
        
        elif field == "XP":
            min_value = self.xp_min_var.get()
            max_value = self.xp_max_var.get()
            if max_value < min_value:
                self.error_label.config(text=f"XP: Maximum value must be bigger or equal than minimum!")
            else:
                self.error_label.config(text="")
        
class Page6(tk.Frame):
    def __init__(self, parent, pages):
        super().__init__(parent)
        self.pages = pages
        self.configure(background="#F5F5DC")

        # Caminhos das imagens
        carta_path = Path(__file__).parent / "imagens/carta.png"

        # Carregar imagem
        if carta_path.exists():
            self.carta_image = tk.PhotoImage(file=str(carta_path))
        else:
            raise FileNotFoundError(f"Imagem '{carta_path}' não encontrada!")

        # Canvas para gerenciar a imagem
        self.canvas = tk.Canvas(self, bg="#F5F5DC", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)

        # Adicionar imagem no canto inferior direito
        self.update_idletasks()  # Atualizar geometria antes de calcular posições
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()
        self.image_id = self.canvas.create_image(
            canvas_width - 10, canvas_height - 10,  # Margem de 10px
            image=self.carta_image,
            anchor="se"  # Âncora no canto inferior direito
        )

        # Redimensionar imagem com a janela
        self.canvas.bind("<Configure>", self.resize_image)

        # Botões na parte superior
        tk.Button(
            self, text="New Combat", font=("Copperplate Gothic Bold", 15),
            bg="lightgreen", command=self.gerar_planilha_excel_com_nome
        ).pack(padx=200, pady=10)

        tk.Button(
            self, text="Add Combat to Campaign", font=("Copperplate Gothic Bold", 15),
            bg="lightblue", command=self.adicionar_combate_arquivo_existente
        ).pack(padx=200, pady=10)

        tk.Button(
            self, text="Campaign Report", font=("Copperplate Gothic Bold", 15),
            bg="lightgrey", command=self.abrir_arquivo_e_gerar_pdf
        ).pack(padx=200, pady=10)

        # Mensagem
        self.message_label = tk.Label(
            self, text="", font=("Copperplate Gothic Light", 12), fg="red", bg="#F5F5DC"
        )
        self.message_label.pack(pady=20)
    
    def resize_image(self, event):
        """
        Ajusta a posição da imagem no canto inferior direito ao redimensionar a janela.
        """
        canvas_width = event.width
        canvas_height = event.height
        self.canvas.coords(self.image_id, canvas_width - 10, canvas_height + 0)
        
    def pegar_textos_das_listas(self):
        # Pega as armas da página 4
        page4 = self.pages[3]
        weapons_content = page4.weapons.get("1.0", "end").strip()
        cleaned_weapons = ",".join(filter(None, weapons_content.split(",")))
        weapon_list = [item.strip() for item in cleaned_weapons.split(",") if item.strip()]

        # Adiciona valor padrão se a lista estiver vazia
        if not weapon_list:
            weapon_list = [" "]

        # Pega os itens da página 5
        page5 = self.pages[4]
        items_content = page5.itens.get("1.0", "end").strip()
        cleaned_items = ",".join(filter(None, items_content.split(",")))
        item_list = [item.strip() for item in cleaned_items.split(",") if item.strip()]

        # Adiciona valor padrão se a lista estiver vazia
        if not item_list:
            item_list = [" "]

        return weapon_list, item_list

    def obter_valores_formulario(self):
        page1, page2, page3, page4, page5, page6 = [self.pages[i] for i in range(6)]
        try:
            # PAGE 1
            battle_name = page1.battle_name_var.get().strip()
            if not battle_name:
                raise ValueError("Battle name is required.")
        
            mob_name = page1.mob_name_var.get().strip()
            quantity = int(page1.quantity_var.get())

            # PAGE 2
            hp_min = int(page2.hp_min_var.get())
            hp_max = int(page2.hp_max_var.get())
            def_min = int(page2.defe_min_var.get())
            def_max = int(page2.defe_max_var.get())

            # PAGE 3
            for_min = int(page3.for_min_var.get())
            for_max = int(page3.for_max_var.get())
            des_min = int(page3.des_min_var.get())
            des_max = int(page3.des_max_var.get())
            con_min = int(page3.con_min_var.get())
            con_max = int(page3.con_max_var.get())
            int_min = int(page3.int_min_var.get())
            int_max = int(page3.int_max_var.get())
            sab_min = int(page3.sab_min_var.get())
            sab_max = int(page3.sab_max_var.get())
            car_min = int(page3.car_min_var.get())
            car_max = int(page3.car_max_var.get())
            dado_init = int(page3.dado_init_value.get())
            modificador = int(page3.mod_init_value.get())

            # PAGE 5
            gold_min = int(page5.gold_min_var.get())
            gold_max = int(page5.gold_max_var.get())
            xp_min = int(page5.xp_min_var.get())
            xp_max = int(page5.xp_max_var.get())
            armas_list, itens_list = self.pegar_textos_das_listas()

            # Validações
            if hp_min > hp_max or def_min > def_max or gold_min > gold_max or xp_min > xp_max:
                raise ValueError("Ensure min values are not greater than max values!")
            if not battle_name or not mob_name or quantity <= 0:
                raise ValueError("Fill all fields in Page 1 correctly!")
            if (
                for_min > for_max or des_min > des_max or con_min > con_max or int_min > int_max or
                sab_min > sab_max or car_min > car_max
            ):
                raise ValueError("Attributes Min cannot exceed Max!")
        
            return {
                "battle_name": battle_name,
                "mob_name": mob_name,
                "quantity": quantity,
                "hp_min": hp_min, "hp_max": hp_max,
                "def_min": def_min, "def_max": def_max,
                "for_min": for_min, "for_max": for_max,
                "des_min": des_min, "des_max": des_max,
                "con_min": con_min, "con_max": con_max,
                "int_min": int_min, "int_max": int_max,
                "sab_min": sab_min, "sab_max": sab_max,
                "car_min": car_min, "car_max": car_max,
                "dado_init": dado_init,
                "modificador": modificador,
                "gold_min": gold_min, "gold_max": gold_max,
                "xp_min": xp_min, "xp_max": xp_max,
                "armas_list": armas_list,
                "itens_list": itens_list
            }
        except ValueError as e:
            self.message_label.config(text=f"Error: {e}", fg="red")
            return None

    def gerar_planilha_excel_com_nome(self):
        valores = self.obter_valores_formulario()
        if not valores:
            return  # Interrompe se houver erro

        
        # PAGE 1
        battle_name = valores["battle_name"]
    
        mob_name = valores["mob_name"]
        quantity = valores["quantity"]

        # PAGE 2
        hp_min = valores["hp_min"]
        hp_max = valores["hp_max"]
        def_min = valores["def_min"]
        def_max = valores["def_max"]

        # PAGE 3
        for_min = valores["for_min"]
        for_max = valores["for_max"]
        des_min = valores["des_min"]
        des_max = valores["des_max"]
        con_min = valores["con_min"]
        con_max = valores["con_max"]
        int_min = valores["int_min"]
        int_max = valores["int_max"]
        sab_min = valores["sab_min"]
        sab_max = valores["sab_max"]
        car_min = valores["car_min"]
        car_max = valores["car_max"]
        dado_init = valores["dado_init"]
        modificador = valores["modificador"]

        # PAGE 5
        gold_min = valores["gold_min"]
        gold_max = valores["gold_max"]
        xp_min = valores["xp_min"]
        xp_max = valores["xp_max"]
    
        armas_list, itens_list = self.pegar_textos_das_listas()



        # Validações
        if hp_min > hp_max or def_min > def_max or gold_min > gold_max or xp_min > xp_max:
            self.message_label.config(text="Error: Ensure min values are not greater than max values!", fg="red")
            return

        if not battle_name or not mob_name or quantity <= 0:
            self.message_label.config(text="Error: Fill all fields in Page 1 correctly!", fg="red")
            return

        if (
            for_min > for_max or des_min > des_max or con_min > con_max or int_min > int_max or
            sab_min > sab_max or car_min > car_max
        ):
            self.message_label.config(text="Error: Attributes Min cannot exceed Max!", fg="red")
            return

        # Criar planilha
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = battle_name

        # Título Principal
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
        cell = ws.cell(row=1, column=1, value=battle_name)
        cell.font = Font(color="000000", bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="C299C2", end_color="C299C2", fill_type="solid")
        cell.border = Border(left=Side(style="thick"),right=Side(style="thick"),top=Side(style="thick"),bottom=Side(style="thick"))
        
        # Weapons Header
        ws.merge_cells(start_row=1, start_column=17, end_row=1, end_column=19)  # Células Q, R, S
        cell = ws.cell(row=1, column=17, value="Weapons")
        cell.font = Font(color="FFFFFF", bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
        cell.border = Border(
            left=Side(style="thick"), right=Side(style="thick"),
            top=Side(style="thick"), bottom=Side(style="thick")
        )

        # Weapons Sub-Headers
        weapon_headers = ["Name", "Damage", "Properties"]
        for col, header in enumerate(weapon_headers, start=17):  # Começa na coluna Q (17)
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="A9CCE3", end_color="A9CCE3", fill_type="solid")
            cell.border = Border(
                left=Side(style="medium"), right=Side(style="medium"),
                top=Side(style="medium"), bottom=Side(style="medium")
            )

        # Weapons Data
        unique_weapons = sorted(set(armas_list))  # Remove duplicatas e ordena as armas
        for row, weapon in enumerate(unique_weapons, start=3):  # Começa na linha 3
            ws.cell(row=row, column=17, value=weapon)  # Nome da arma
            ws.cell(row=row, column=18, value="1d")  # Substituir por valor real
            ws.cell(row=row, column=19, value=" ")  # Substituir por valor real
            ws.column_dimensions['S'].width = 25
            ws.column_dimensions['Q'].width = 10

            # Alternar cores entre dois tons de azul
            color = "D6EAF8" if row % 2 == 0 else "A9CCE3"
            for col in range(17, 20):  # Colunas Q, R, S
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.border = Border(
                    left=Side(style="medium"), right=Side(style="medium"),
                    top=Side(style="medium"), bottom=Side(style="medium")
                )

            # Adicionar linha separadora
            if row != len(unique_weapons) + 2:  # Evitar borda extra na última linha
                for col in range(17, 20):  # Colunas Q, R, S
                    ws.cell(row=row, column=col).border = Border(
                        left=Side(style="medium"), right=Side(style="medium"),
                        top=Side(style="medium"), bottom=Side(style="thin")  # Linha inferior fina
                    )

            
        # Linha em branco após Weapons
        row += 1  # Adicionar uma linha em branco

        # Players Header
        row += 1  # Avança para a próxima linha
        ws.cell(row=row, column=17, value="Players")
        ws.merge_cells(start_row=row, start_column=17, end_row=row, end_column=19)
        cell = ws.cell(row=row, column=17)
        cell.font = Font(color="000000", bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
        cell.border = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="medium"), bottom=Side(style="medium")
        )

        # Players Sub-Headers
        row += 1
        player_headers = ["Name", "Defense", "Initiative"]
        for col, header in enumerate(player_headers, start=17):  # Colunas Q, R, S
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
            cell.border = Border(
                left=Side(style="medium"), right=Side(style="medium"),
                top=Side(style="medium"), bottom=Side(style="medium")
            )

        # Players Data with alternating row colors
        for i in range(5):  # Criar 5 linhas para os jogadores
            row += 1
            ws.cell(row=row, column=17, value=f"Player {i + 1}")  # Nome do jogador
            ws.cell(row=row, column=18, value=" ")  # Defesa
            ws.cell(row=row, column=19, value=" ")  # Iniciativa
    
            # Alternar cores entre dois tons de amarelo
            color = "FEF9E7" if row % 2 == 0 else "FDF5E6"
            for col in range(17, 20):  # Colunas Q, R, S
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.border = Border(
                    left=Side(style="medium"), right=Side(style="medium"),
                    top=Side(style="medium"), bottom=Side(style="medium")
                )



        # Cabeçalhos
        headers = [
            "Name", "Initiative", "Defense", "HP", "Damage", "Attack",
            "Str", "Dex", "Con", "Int", "Wis", "Cha",
            "Gold", "XP", "Items"
        ]
        header_colors = [
            "A0522D", "87CEEB", "808080", "006400", "FF0000", "4682B4",
            "FFA500", "FFA500", "FFA500", "FFA500", "FFA500", "FFA500",
            "FFFF00", "20B2AA", "FA8072"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = PatternFill(start_color=header_colors[col - 1], end_color=header_colors[col - 1], fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="medium"),
                right=Side(style="medium"),
                top=Side(style="medium"),
                bottom=Side(style="medium")
            )
        
        # Aumentar o campo do nome
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["O"].width = 15

        # Estilos das colunas a partir da linha 3
        column_colors = {
            1: "D2B48C",  # Marrom fraco
            2: "ADD8E6",  # Azul céu fraco
            3: "D3D3D3",  # Cinza
            4: "90EE90",  # Verde fraco
            5: "FFB6C1",  # Vermelho fraco
            6: "B0C4DE",  # Branco marfim
            7: "FFDAB9",  # Laranja fraco
            8: "FFDAB9",  # Laranja fraco
            9: "FFDAB9",  # Laranja fraco
            10: "FFDAB9", # Laranja fraco
            11: "FFDAB9", # Laranja fraco
            12: "FFDAB9", # Laranja fraco
            13: "FFFFE0", # Amarelo fraco
            14: "B2DFD8", # Verde água fraco
            15: "FFA07A"  # Salmão fraco
        }

        # Linhas de dados
        for i in range(quantity):
            row = i + 3
            mob_name_with_number = f"{mob_name} {i + 1}"
            ws.cell(row=row, column=1, value=mob_name_with_number)
            
            try:
                # Define um valor padrão para dado_init caso não seja válido
                dado_init = max(dado_init, 1) if isinstance(dado_init, int) else 1
                modificador = modificador if isinstance(modificador, int) else 0  # Verifica se modificador é válido
                initiative = random.randint(1, dado_init) + modificador
                ws.cell(row=row, column=2, value=initiative)
            except Exception as e:
                print(f"Erro ao calcular iniciativa: {e}")
                initiative = 0
                ws.cell(row=row, column=2, value=initiative)


            defense = random.randint(def_min, def_max)
            ws.cell(row=row, column=3, value=defense)

            hp = random.randint(hp_min, hp_max)
            ws.cell(row=row, column=4, value=hp)

            ws.cell(row=row, column=5, value="")

            weapon = random.choice(armas_list)
            ws.cell(row=row, column=6, value=weapon)

            attributes = [
                random.randint(for_min, for_max),
                random.randint(des_min, des_max),
                random.randint(con_min, con_max),
                random.randint(int_min, int_max),
                random.randint(sab_min, sab_max),
                random.randint(car_min, car_max),
            ]
            for col, value in enumerate(attributes, start=7):
                ws.cell(row=row, column=col, value=value)

            # Gold e XP
            gold = random.randint(gold_min, gold_max)
            ws.cell(row=row, column=13, value=gold)

            xp = random.randint(xp_min, xp_max)
            ws.cell(row=row, column=14, value=xp)

            # Itens
            item = random.choice(itens_list)
            ws.cell(row=row, column=15, value=item)

            # Aplicar cores às colunas da linha atual
            for col in range(1, 16):
                cell = ws.cell(row=row, column=col)
                if col in column_colors:
                    color = column_colors[col]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.border = Border(left=Side(style="medium"),right=Side(style="medium"),top=Side(style="medium"),bottom=Side(style="medium"))
                    
        # Adicionar formatação condicional para coluna "A"
        for i in range(3, 3 + quantity):
            cell_reference_A = f"A{i}"
            cell_reference_D = f"D{i}"
            cell_reference_E = f"E{i}"

            # Regra para vermelho (E >= D)
            red_rule = FormulaRule(
                formula=[f"${cell_reference_E} >= ${cell_reference_D}"],
                stopIfTrue=True,
                fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            )

            # Regra para marrom (E <= D)
            brown_fill = PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid")

            ws.conditional_formatting.add(cell_reference_A, red_rule)

            # Define o fundo padrão como marrom
            ws[cell_reference_A].fill = brown_fill

        filename = f"{battle_name}.xlsx"
        wb.save(filename)
        
        # 2. Usar Pandas para calcular as somas
        try:
            # Carregar o Excel com Pandas, pulando as primeiras linhas
            df = pd.read_excel(filename, sheet_name=battle_name, engine="openpyxl", header=None, skiprows=2)

            # Renomear colunas para facilitar o acesso (coluna 'D', 'M', e 'N' específicas)
            df.rename(columns={3: "HP", 12: "Gold", 13: "XP"}, inplace=True)

            # Calcular as somas
            total_gold = df["Gold"].sum()
            total_xp = df["XP"].sum()
            total_hp = df["HP"].sum()

            # Adicionar as somas no Excel
            with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df_totals = pd.DataFrame({
                    "Attribute": ["Total Gold", "Total XP", "Total HP"],
                    "Value": [total_gold, total_xp, total_hp]
                })
                # Escreve a tabela abaixo da já existente
                df_totals.to_excel(writer, sheet_name=battle_name, index=False, startrow=len(df) + 3, startcol=16)

            self.message_label.config(text=f"Combate '{battle_name}' criado e somas calculadas com sucesso!", fg="green")
        except Exception as e:
            self.message_label.config(text=f"Erro ao calcular somas: {e}", fg="red")


    def adicionar_combate_arquivo_existente(self):
        filename = filedialog.askopenfilename(title="Select Campaign File", filetypes=[("Excel files", "*.xlsx")])
        if not filename:
            return

        wb = openpyxl.load_workbook(filename, read_only=False)  # Abrindo o arquivo para edição
        ws_names = wb.sheetnames

        valores = self.obter_valores_formulario()
        if not valores:
            return  # Interrompe se houver erro

        
        # PAGE 1
        battle_name = valores["battle_name"]
    
        mob_name = valores["mob_name"]
        quantity = valores["quantity"]

        # PAGE 2
        hp_min = valores["hp_min"]
        hp_max = valores["hp_max"]
        def_min = valores["def_min"]
        def_max = valores["def_max"]

        # PAGE 3
        for_min = valores["for_min"]
        for_max = valores["for_max"]
        des_min = valores["des_min"]
        des_max = valores["des_max"]
        con_min = valores["con_min"]
        con_max = valores["con_max"]
        int_min = valores["int_min"]
        int_max = valores["int_max"]
        sab_min = valores["sab_min"]
        sab_max = valores["sab_max"]
        car_min = valores["car_min"]
        car_max = valores["car_max"]
        dado_init = valores["dado_init"]
        modificador = valores["modificador"]

        # PAGE 5
        gold_min = valores["gold_min"]
        gold_max = valores["gold_max"]
        xp_min = valores["xp_min"]
        xp_max = valores["xp_max"]
    
        armas_list, itens_list = self.pegar_textos_das_listas()

        if hp_min > hp_max or def_min > def_max or gold_min > gold_max or xp_min > xp_max:
            self.message_label.config(text="Error: Ensure min values are not greater than max values!", fg="red")
            return

        if (
            for_min > for_max or des_min > des_max or con_min > con_max or int_min > int_max or
            sab_min > sab_max or car_min > car_max
        ):
            self.message_label.config(text="Error: Attributes Min cannot exceed Max!", fg="red")
            return

        # Garantir que o nome da aba seja único
        original_name = battle_name
        while battle_name in ws_names:
            battle_name = f"{original_name}_New"
    
        ws = wb.create_sheet(title=battle_name)

        
        # Título Principal
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
        cell = ws.cell(row=1, column=1, value=battle_name)
        cell.font = Font(color="000000", bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="C299C2", end_color="C299C2", fill_type="solid")
        cell.border = Border(left=Side(style="thick"),right=Side(style="thick"),top=Side(style="thick"),bottom=Side(style="thick"))
        
        # Weapons Header
        ws.merge_cells(start_row=1, start_column=17, end_row=1, end_column=19)  # Células Q, R, S
        cell = ws.cell(row=1, column=17, value="Weapons")
        cell.font = Font(color="FFFFFF", bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
        cell.border = Border(
            left=Side(style="thick"), right=Side(style="thick"),
            top=Side(style="thick"), bottom=Side(style="thick")
        )

        # Weapons Sub-Headers
        weapon_headers = ["Name", "Damage", "Properties"]
        for col, header in enumerate(weapon_headers, start=17):  # Começa na coluna Q (17)
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="A9CCE3", end_color="A9CCE3", fill_type="solid")
            cell.border = Border(
                left=Side(style="medium"), right=Side(style="medium"),
                top=Side(style="medium"), bottom=Side(style="medium")
            )

        # Weapons Data
        unique_weapons = sorted(set(armas_list))  # Remove duplicatas e ordena as armas
        for row, weapon in enumerate(unique_weapons, start=3):  # Começa na linha 3
            ws.cell(row=row, column=17, value=weapon)  # Nome da arma
            ws.cell(row=row, column=18, value="1d")  # Substituir por valor real
            ws.cell(row=row, column=19, value=" ")  # Substituir por valor real
            ws.column_dimensions['S'].width = 25
            ws.column_dimensions['Q'].width = 10

            # Alternar cores entre dois tons de azul
            color = "D6EAF8" if row % 2 == 0 else "A9CCE3"
            for col in range(17, 20):  # Colunas Q, R, S
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.border = Border(
                    left=Side(style="medium"), right=Side(style="medium"),
                    top=Side(style="medium"), bottom=Side(style="medium")
                )

            # Adicionar linha separadora
            if row != len(unique_weapons) + 2:  # Evitar borda extra na última linha
                for col in range(17, 20):  # Colunas Q, R, S
                    ws.cell(row=row, column=col).border = Border(
                        left=Side(style="medium"), right=Side(style="medium"),
                        top=Side(style="medium"), bottom=Side(style="thin")  # Linha inferior fina
                    )

            
        # Linha em branco após Weapons
        row += 1  # Adicionar uma linha em branco

        # Players Header
        row += 1  # Avança para a próxima linha
        ws.cell(row=row, column=17, value="Players")
        ws.merge_cells(start_row=row, start_column=17, end_row=row, end_column=19)
        cell = ws.cell(row=row, column=17)
        cell.font = Font(color="000000", bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
        cell.border = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="medium"), bottom=Side(style="medium")
        )

        # Players Sub-Headers
        row += 1
        player_headers = ["Name", "Defense", "Initiative"]
        for col, header in enumerate(player_headers, start=17):  # Colunas Q, R, S
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
            cell.border = Border(
                left=Side(style="medium"), right=Side(style="medium"),
                top=Side(style="medium"), bottom=Side(style="medium")
            )

        # Players Data with alternating row colors
        for i in range(5):  # Criar 5 linhas para os jogadores
            row += 1
            ws.cell(row=row, column=17, value=f"Player {i + 1}")  # Nome do jogador
            ws.cell(row=row, column=18, value=" ")  # Defesa
            ws.cell(row=row, column=19, value=" ")  # Iniciativa
    
            # Alternar cores entre dois tons de amarelo
            color = "FEF9E7" if row % 2 == 0 else "FDF5E6"
            for col in range(17, 20):  # Colunas Q, R, S
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.border = Border(
                    left=Side(style="medium"), right=Side(style="medium"),
                    top=Side(style="medium"), bottom=Side(style="medium")
                )
        

        # Cabeçalhos
        headers = [
            "Name", "Initiative", "Defense", "HP", "Damage", "Attack",
            "Str", "Dex", "Con", "Int", "Wis", "Cha",
            "Gold", "XP", "Items"
        ]
        header_colors = [
            "A0522D", "87CEEB", "808080", "006400", "FF0000", "4682B4",
            "FFA500", "FFA500", "FFA500", "FFA500", "FFA500", "FFA500",
            "FFFF00", "20B2AA", "FA8072"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = PatternFill(start_color=header_colors[col - 1], end_color=header_colors[col - 1], fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="medium"),
                right=Side(style="medium"),
                top=Side(style="medium"),
                bottom=Side(style="medium")
            )
        
        # Aumentar o campo do nome
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["O"].width = 15

        # Estilos das colunas a partir da linha 3
        column_colors = {
            1: "D2B48C",  # Marrom fraco
            2: "ADD8E6",  # Azul céu fraco
            3: "D3D3D3",  # Cinza
            4: "90EE90",  # Verde fraco
            5: "FFB6C1",  # Vermelho fraco
            6: "B0C4DE",  # Branco marfim
            7: "FFDAB9",  # Laranja fraco
            8: "FFDAB9",  # Laranja fraco
            9: "FFDAB9",  # Laranja fraco
            10: "FFDAB9", # Laranja fraco
            11: "FFDAB9", # Laranja fraco
            12: "FFDAB9", # Laranja fraco
            13: "FFFFE0", # Amarelo fraco
            14: "B2DFD8", # Verde água fraco
            15: "FFA07A"  # Salmão fraco
        }

        # Linhas de dados
        for i in range(quantity):
            row = i + 3
            mob_name_with_number = f"{mob_name} {i + 1}"
            ws.cell(row=row, column=1, value=mob_name_with_number)

            try:
                # Define um valor padrão para dado_init caso não seja válido
                dado_init = max(dado_init, 1) if isinstance(dado_init, int) else 1
                modificador = modificador if isinstance(modificador, int) else 0  # Verifica se modificador é válido
                initiative = random.randint(1, dado_init) + modificador
                ws.cell(row=row, column=2, value=initiative)
            except Exception as e:
                self.message_label.config(text=f"Error em calcular a iniciativa: {e}", fg="red")
                initiative = 0
                ws.cell(row=row, column=2, value=initiative)

            defense = random.randint(def_min, def_max)
            ws.cell(row=row, column=3, value=defense)

            hp = random.randint(hp_min, hp_max)
            ws.cell(row=row, column=4, value=hp)

            ws.cell(row=row, column=5, value="")

            weapon = random.choice(armas_list)
            ws.cell(row=row, column=6, value=weapon)

            attributes = [
                random.randint(for_min, for_max),
                random.randint(des_min, des_max),
                random.randint(con_min, con_max),
                random.randint(int_min, int_max),
                random.randint(sab_min, sab_max),
                random.randint(car_min, car_max),
            ]
            for col, value in enumerate(attributes, start=7):
                ws.cell(row=row, column=col, value=value)

            # Gold e XP
            gold = random.randint(gold_min, gold_max)
            ws.cell(row=row, column=13, value=gold)

            xp = random.randint(xp_min, xp_max)
            ws.cell(row=row, column=14, value=xp)

            # Itens
            item = random.choice(itens_list)
            ws.cell(row=row, column=15, value=item)

            # Aplicar cores às colunas da linha atual
            for col in range(1, 16):
                cell = ws.cell(row=row, column=col)
                if col in column_colors:
                    color = column_colors[col]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.border = Border(left=Side(style="medium"),right=Side(style="medium"),top=Side(style="medium"),bottom=Side(style="medium"))
                    
        # Adicionar formatação condicional para coluna "A"
        for i in range(3, 3 + quantity):
            cell_reference_A = f"A{i}"
            cell_reference_D = f"D{i}"
            cell_reference_E = f"E{i}"

            # Regra para vermelho (E >= D)
            red_rule = FormulaRule(
                formula=[f"${cell_reference_E} >= ${cell_reference_D}"],
                stopIfTrue=True,
                fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            )

            # Regra para marrom (E <= D)
            brown_fill = PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid")

            ws.conditional_formatting.add(cell_reference_A, red_rule)

            # Define o fundo padrão como marrom
            ws[cell_reference_A].fill = brown_fill

        
        wb.save(filename)
        
        # 2. Usar Pandas para calcular as somas
        try:
            # Carregar o Excel com Pandas, pulando as primeiras linhas
            df = pd.read_excel(filename, sheet_name=battle_name, engine="openpyxl", header=None, skiprows=2)

            # Renomear colunas para facilitar o acesso (coluna 'D', 'M', e 'N' específicas)
            df.rename(columns={3: "HP", 12: "Gold", 13: "XP"}, inplace=True)

            # Calcular as somas
            total_gold = df["Gold"].sum()
            total_xp = df["XP"].sum()
            total_hp = df["HP"].sum()

            # Adicionar as somas no Excel
            with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df_totals = pd.DataFrame({
                    "Attribute": ["Total Gold", "Total XP", "Total HP"],
                    "Value": [total_gold, total_xp, total_hp]
                })
                # Escreve a tabela abaixo da já existente
                df_totals.to_excel(writer, sheet_name=battle_name, index=False, startrow=len(df) + 3, startcol=16)

            self.message_label.config(text=f"Combate '{battle_name}' created successfully!", fg="green")
        except Exception as e:
            self.message_label.config(text=f"Error at sum: {e}", fg="red")

    def abrir_arquivo_e_gerar_pdf(self):
        # Configura a janela do Tkinter
        root = tk.Tk()
        root.withdraw()  # Oculta a janela principal

        # Caixa de diálogo para selecionar o arquivo Excel
        excel_path = filedialog.askopenfilename(
            title="Select a Excel File",
            filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*"))
        )

        if not excel_path:
            root.destroy()
            self.message_label.config(text="Error: No file was selected.", fg="red")
            return

        # Caixa de diálogo para salvar o arquivo PDF
        pdf_path = filedialog.asksaveasfilename(
            title="Save as PDF",
            defaultextension=".pdf",
            filetypes=(("Arquivos PDF", "*.pdf"), ("Todos os arquivos", "*.*"))
        )
        
        root.destroy()

        if not pdf_path:
            self.message_label.config(text="Error: No path was selected!", fg="red")
            return

        # Gera o relatório
        self.gerar_relatorio_em_pdf(excel_path, pdf_path)
        self.message_label.config(text=f"Report save in: {pdf_path}", fg="green")

    def gerar_relatorio_em_pdf(self, excel_path, pdf_path):
        
        try:
            
            xlsx = pd.ExcelFile(excel_path)
            
            # Cabeçalhos esperados
            expected_headers = {"Attack", "HP", "Gold", "XP", "Items"}
            headers_found = set()
            
            # Verificar os cabeçalhos em todas as abas
            for sheet_name in xlsx.sheet_names:
                df = pd.read_excel(xlsx, sheet_name=sheet_name, skiprows=1)
                headers_found.update(df.columns)
                
                if expected_headers.issubset(headers_found):
                    break
                
            # Verificar se todos os cabeçalhos esperados estão presentes
            missing_headers = expected_headers - headers_found
            if missing_headers:
                self.message_label.config(
                    text=f"Error: The file is invalid. Missing headers: {', '.join(missing_headers)}",
                    fg="red"
                )
                return

            total_gold = 0
            total_xp = 0
            total_hp = 0
            total_attack = Counter()
            total_enemies = 0
            aba_xp_gold = {}

            # Primeiro gráfico: XP por aba
            plt.figure(figsize=(10, 6))
            xp_totals = []
            for aba in xlsx.sheet_names:
                df = pd.read_excel(xlsx, sheet_name=aba, skiprows=1)
                df['XP'] = pd.to_numeric(df['XP'], errors='coerce').fillna(0)
                xp_sum = df['XP'].sum()
                xp_totals.append(xp_sum)
                total_xp += xp_sum
            plt.bar(xlsx.sheet_names, xp_totals, color='blue')
            plt.title("XP per Battle")
            plt.xlabel("Battle")
            plt.ylabel("XP Total")
            xp_chart_path = "xp_chart.png"
            plt.savefig(xp_chart_path)
            plt.close()

            # Segundo gráfico: Gold por aba
            plt.figure(figsize=(10, 6))
            gold_totals = []
            for aba in xlsx.sheet_names:
                df = pd.read_excel(xlsx, sheet_name=aba, skiprows=1)
                df['Gold'] = pd.to_numeric(df['Gold'], errors='coerce').fillna(0)
                gold_sum = df['Gold'].sum()
                gold_totals.append(gold_sum)
                total_gold += gold_sum
            plt.bar(xlsx.sheet_names, gold_totals, color='green')
            plt.title("Gold per Battle")
            plt.xlabel("Battle")
            plt.ylabel("Gold Total")
            gold_chart_path = "gold_chart.png"
            plt.savefig(gold_chart_path)
            plt.close()

            # PDF Canvas
            c = canvas.Canvas(pdf_path, pagesize=letter)
            width, height = letter

            # Página 1: Gráficos de XP e Gold por Aba
            c.drawImage(xp_chart_path, 50, height / 2 + 50, width=500, height=300)
            c.drawImage(gold_chart_path, 50, 50, width=500, height=300)
            c.showPage()

            # Contagem total de ataques
            for aba in xlsx.sheet_names:
                df = pd.read_excel(xlsx, sheet_name=aba, skiprows=1)
                df['HP'] = pd.to_numeric(df['HP'], errors='coerce').fillna(0)
                total_hp += df['HP'].sum()
                df['Attack'] = df['Attack'].dropna().astype(str).str.strip()
                filtered_attacks = df['Attack'][df['Attack'] != ""]  # Filtra células vazias ou "nan"
                total_attack.update(Counter(filtered_attacks))

                # Contar inimigos preenchidos na coluna A a partir da linha 3
                total_enemies += df.iloc[:, 0].dropna().shape[0]

            # Gráfico de pizza: Frequência de ataques
            plt.figure(figsize=(8, 8))
            # Filtrar valores que não sejam "nan" e tenham ocorrências válidas
            filtered_attacks = {key: value for key, value in total_attack.items() if pd.notna(key) and value > 0}
            plt.pie(filtered_attacks.values(), labels=filtered_attacks.keys(), autopct='%1.1f%%', startangle=140)
            plt.title("Weapon Frequency")
            attack_chart_path = "attack_chart.png"
            plt.savefig(attack_chart_path)
            plt.close()

            # Página 2: Gráfico de pizza e detalhes
            c.drawImage(attack_chart_path, 50, height / 2, width=500, height=300)

            # Adicionar detalhes abaixo do gráfico
            c.setFont("Helvetica-Bold", 14)
            c.drawString(50, height / 2 - 50, "Details")
            c.setFont("Helvetica", 12)
            c.drawString(50, height / 2 - 80, f"Total Gold: {total_gold}")
            c.drawString(50, height / 2 - 100, f"Total XP: {total_xp}")
            c.drawString(50, height / 2 - 120, f"Total Damage: {total_hp}")
            c.drawString(50, height / 2 - 140, f"Total Enemies: {total_enemies}")

            c.save()

            # Remover arquivos temporários
            os.remove(xp_chart_path)
            os.remove(gold_chart_path)
            os.remove(attack_chart_path)
        
        except Exception as e:
            # Tratar erros inesperados
            self.message_label.config(text=f"Unexpected error occurred: {str(e)}",fg="red")


        
if __name__ == "__main__":
    Mobster().mainloop()